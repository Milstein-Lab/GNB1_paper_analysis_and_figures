"""
generate_sex_stratified_figures.py
===================================
Generates sex-stratified (Male vs Female) figures and statistics for the GNB1 manuscript.
For each metric, plots are broken down by Genotype × Sex (4 groups: WT-M, WT-F, I80T/+-M, I80T/+-F).

Metrics covered:
  1. Summed Activity – All Dark Hours (DVC)
  2. FI Midpoint
  3. E/I Imbalance – All three pathways (Perforant, Schaffer, Basal_Stratum_Oriens)
  4. Plateau Area
  5. Open Field Locomotion (Distance) & Open Field Anxiety (Center:Outer Ratio)
  6. OLM (Object Location Memory) – Testing Discrimination Index

Stats (Mann-Whitney U) are computed for:
  - Male:   WT-M vs I80T/+-M
  - Female: WT-F vs I80T/+-F
Results are appended to paper_data/Master_Stats_Summary.csv and .xlsx as new tabs.

Figures are saved to paper_figures/ as .png and .svg.
"""

import os
import sys
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from scipy import stats

# Import plotting utilities (same style as generate_figures.py)
from plotting_utils import (
    setup_publication_style,
    save_current_fig,
    rename_genotype,
    apply_clean_yticks,
    get_safe_y,
    add_subplot_label,
    draw_significance,
    COLORS,
)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
PAPER_DATA_DIR  = "paper_data"
OUTPUT_FIG_DIR  = "paper_figures"
MASTER_CSV      = os.path.join(PAPER_DATA_DIR, "Master_Stats_Summary.csv")
MASTER_XLSX     = os.path.join(PAPER_DATA_DIR, "Master_Stats_Summary.xlsx")

# Genotype × Sex groups
# Colors: WT=black shades, I80T/+=red shades;  Male=lighter, Female=darker
SEX_GENO_ORDER  = ["WT-M", "WT-F", "I80T/+-M", "I80T/+-F"]

# Okabe-Ito colorblind-safe palette
# WT   → blue family:    sky-blue (Male lighter) / dark-blue (Female darker)
# I80T → orange family:  orange   (Male lighter) / vermilion (Female darker)
# Linestyle also encodes sex: dashed = Male, solid = Female
SEX_GENO_COLORS = {
    "WT-M":     "#56B4E9",   # sky blue
    "WT-F":     "#0072B2",   # dark blue
    "I80T/+-M": "#E69F00",   # orange
    "I80T/+-F": "#D55E00",   # vermilion
}
SEX_GENO_LINES = {           # linestyle by sex (reinforces color)
    "WT-M":     "--",
    "WT-F":     "-",
    "I80T/+-M": "--",
    "I80T/+-F": "-",
}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def p_to_sig(p):
    try:
        p = float(p)
    except (ValueError, TypeError):
        return "?"
    if p < 0.001:   return "***"
    elif p < 0.01:  return "**"
    elif p < 0.05:  return "*"
    else:           return "ns"


def mann_whitney(a, b):
    """Returns (U, p). Returns (nan, nan) if either group is too small."""
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    a = a[~np.isnan(a)]
    b = b[~np.isnan(b)]
    if len(a) < 2 or len(b) < 2:
        return np.nan, np.nan
    u, p = stats.mannwhitneyu(a, b, alternative="two-sided")
    return u, p


def make_sex_geno_col(df, genotype_col="Genotype", sex_col="Sex"):
    """Creates a combined Genotype-Sex label column for grouping."""
    df = df.copy()
    # Normalize sex: M/Male -> M, F/Female -> F
    sex_map = {"M": "M", "Male": "M", "male": "M",
               "F": "F", "Female": "F", "female": "F",
               "Male ": "M"}  # strip trailing spaces
    df["_Sex"]  = df[sex_col].map(sex_map)
    df["_Geno"] = df[genotype_col]
    df["SexGeno"] = df["_Geno"] + "-" + df["_Sex"]
    return df


def plot_sex_bar_scatter(ax, data, y_col, title, ylabel,
                         order=None, show_pval_bracket=True):
    """
    Bar + scatter plot for 4 sex×genotype groups.
    Returns a list of stats dicts: [{group_comparison, U, p, sig}, ...].
    """
    if order is None:
        order = SEX_GENO_ORDER
    color_map = SEX_GENO_COLORS

    bar_width = 0.55
    stats_out = []

    # Marker shape encodes sex: Male = triangle, Female = circle
    MARKER_MAP = {"WT-M": "^", "WT-F": "o", "I80T/+-M": "^", "I80T/+-F": "o"}

    for i, group in enumerate(order):
        subset = data[data["SexGeno"] == group]
        if subset.empty:
            continue
        values = subset[y_col].dropna().values
        if len(values) == 0:
            continue
        color  = color_map.get(group, "gray")
        marker = MARKER_MAP.get(group, "o")
        mean   = np.mean(values)
        sem    = np.std(values, ddof=1) / np.sqrt(len(values)) if len(values) > 1 else 0

        ax.bar(i, mean, width=bar_width, color=color, alpha=0.55, edgecolor="none")
        ax.errorbar(i, mean, yerr=sem, fmt="o", color=color,
                    capsize=1, elinewidth=1, markersize=2)
        # Jittered scatter with sex-encoded marker
        jitter = np.random.default_rng(42).uniform(-0.12, 0.12, len(values))
        ax.scatter(np.full(len(values), i) + jitter, values,
                   color=color, marker=marker, s=10, zorder=3, alpha=0.85,
                   linewidths=0.4, edgecolors="white")

    ax.set_xticks(range(len(order)))
    labels = []
    for g in order:
        sub = data[data["SexGeno"] == g].dropna(subset=[y_col])
        n   = len(sub)
        labels.append(f"{g}\n(n={n})")
    ax.set_xticklabels(labels, fontsize=6)
    ax.set_ylabel(ylabel, fontsize=7)
    ax.set_title(title, fontsize=8)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    # Smart y-limits
    all_vals = data[y_col].dropna().values
    if len(all_vals):
        lo, hi = all_vals.min(), all_vals.max()
        rng = hi - lo if hi != lo else abs(hi) or 1
        if lo >= 0:
            ax.set_ylim(0, hi + rng * 0.25)
        elif hi <= 0:
            ax.set_ylim(lo - rng * 0.25, 0)
        else:
            ax.set_ylim(lo - rng * 0.2, hi + rng * 0.25)
    apply_clean_yticks(ax)

    # ── Significance brackets ──────────────────────────────────────────────
    if show_pval_bracket:
        ylo, yhi = ax.get_ylim()
        y_range  = yhi - ylo

        # Male comparison: WT-M (idx=0) vs I80T/+-M (idx=2)
        wt_m  = data[data["SexGeno"] == "WT-M"][y_col].dropna().values
        mut_m = data[data["SexGeno"] == "I80T/+-M"][y_col].dropna().values
        u_m, p_m = mann_whitney(wt_m, mut_m)
        if not np.isnan(p_m):
            y_br_m = yhi - y_range * 0.05
            draw_significance(ax, 0, 2, p_m, y_br_m, bracket=True)
            stats_out.append(dict(Comparison="WT-M vs I80T/+-M", U=u_m, P=p_m, Sig=p_to_sig(p_m)))

        # Female comparison: WT-F (idx=1) vs I80T/+-F (idx=3)
        wt_f  = data[data["SexGeno"] == "WT-F"][y_col].dropna().values
        mut_f = data[data["SexGeno"] == "I80T/+-F"][y_col].dropna().values
        u_f, p_f = mann_whitney(wt_f, mut_f)
        if not np.isnan(p_f):
            y_br_f = yhi - y_range * 0.12
            draw_significance(ax, 1, 3, p_f, y_br_f, bracket=True)
            stats_out.append(dict(Comparison="WT-F vs I80T/+-F", U=u_f, P=p_f, Sig=p_to_sig(p_f)))

    return stats_out


def save_fig_png_svg(fig, fig_name):
    """Saves the current figure as both .png and .svg to paper_figures/."""
    os.makedirs(OUTPUT_FIG_DIR, exist_ok=True)
    base = os.path.join(OUTPUT_FIG_DIR, fig_name)
    fig.savefig(base + ".png", dpi=300, bbox_inches="tight")
    fig.savefig(base + ".svg", bbox_inches="tight", format="svg",
                transparent=False, dpi=300)
    print(f"  ✓ Saved {fig_name}.png  +  .svg")
    plt.close(fig)


def build_stats_row(figure, subpanel, metric, pathway, condition,
                    data, y_col, sex_label, comparison):
    """
    Build a stats dict in Master_Stats_Summary format for one sex comparison.
    Comparison is e.g. 'WT-M vs I80T/+-M'.
    """
    geno_a, geno_b = comparison.split(" vs ")
    grp_a = data[data["SexGeno"] == geno_a][y_col].dropna()
    grp_b = data[data["SexGeno"] == geno_b][y_col].dropna()
    u, p  = mann_whitney(grp_a.values, grp_b.values)

    def _f(x): return round(float(x), 4) if pd.notna(x) and not np.isnan(float(x)) else np.nan
    def _i(x): return int(x) if pd.notna(x) and len(x) > 0 else np.nan

    return dict(
        Figure=figure, Subpanel=subpanel, Metric=metric + f" [{sex_label}]",
        Pathway=pathway, Condition=condition,
        WT_Mean=_f(grp_a.mean()), WT_SEM=_f(grp_a.sem()),
        WT_N=len(grp_a),
        I80T_Mean=_f(grp_b.mean()), I80T_SEM=_f(grp_b.sem()),
        I80T_N=len(grp_b),
        Test_Used="Mann-Whitney U (sex-stratified)",
        Statistic=_f(u),
        P_Value=_f(p),
        Significance=p_to_sig(p),
        Notes=f"Sex-stratified: {comparison}",
    )


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_all_data():
    """Load all required datasets and attach sex info where needed."""
    data = {}

    # master_df (contains Cell_ID and Sex for physiology cells)
    master = pd.read_csv("master_df.csv", low_memory=False)
    master["Cell_ID"] = master["Cell_ID"].astype(str).str.strip()
    data["master"] = master

    # 1. DVC – Summed Activity (All Dark Hours)
    df_cage = pd.read_csv(os.path.join(PAPER_DATA_DIR, "DVC_Analysis",
                                        "Cage_Specific_Hours_Summary.csv"))
    df_cage = rename_genotype(df_cage)
    df_cage = make_sex_geno_col(df_cage, sex_col="Sex")
    data["dvc_cage"] = df_cage

    # 2. FI Midpoint – join sex from master_df
    df_fi = pd.read_csv(os.path.join(PAPER_DATA_DIR, "Firing_Rate",
                                      "Sigmoid_Fit_Params.csv"))
    df_fi["Cell_ID"] = df_fi["Cell_ID"].astype(str).str.strip()
    # Merge sex from master_df
    sex_map = master[["Cell_ID", "Sex"]].drop_duplicates()
    df_fi = df_fi.merge(sex_map, on="Cell_ID", how="left")
    df_fi = df_fi.rename(columns={"Midpoint": "FI_Midpoint"}) \
                  if "FI_Midpoint" not in df_fi.columns else df_fi
    df_fi = rename_genotype(df_fi)
    df_fi = make_sex_geno_col(df_fi, sex_col="Sex")
    data["fi_midpoint"] = df_fi

    # 3. E/I Imbalance – already has Sex col in E_I_amplitudes.csv
    df_ei = pd.read_csv(os.path.join(PAPER_DATA_DIR, "E_I_data",
                                      "E_I_amplitudes.csv"))
    df_ei = rename_genotype(df_ei)
    df_ei = make_sex_geno_col(df_ei, sex_col="Sex")
    data["ei"] = df_ei

    # 4. Plateau Area
    df_plat = pd.read_csv(os.path.join(PAPER_DATA_DIR, "Plateau_data",
                                         "Plateau_data.csv"))
    df_plat = rename_genotype(df_plat)
    # Normalize sex labels ("Male " -> "M", "Female" -> "F")
    if "Sex" in df_plat.columns:
        df_plat = make_sex_geno_col(df_plat, sex_col="Sex")
    data["plateau"] = df_plat

    # 5. Open Field Locomotion & Anxiety
    df_of_loc = pd.read_csv(os.path.join(PAPER_DATA_DIR, "Behavior_Analysis",
                                           "Open_Field_Locomotion_Trial1.csv"))
    df_of_anx = pd.read_csv(os.path.join(PAPER_DATA_DIR, "Behavior_Analysis",
                                           "Open_Field_Anxiety_Processed.csv"))
    df_of_loc = rename_genotype(df_of_loc)
    df_of_anx = rename_genotype(df_of_anx)
    df_of_loc = make_sex_geno_col(df_of_loc, sex_col="Sex")
    df_of_anx = make_sex_geno_col(df_of_anx, sex_col="Sex")
    data["of_loc"] = df_of_loc
    data["of_anx"] = df_of_anx

    # 6. OLM
    df_olm = pd.read_csv(os.path.join(PAPER_DATA_DIR, "Behavior_Analysis",
                                        "OLM_Summary_Deltas.csv"))
    df_olm = rename_genotype(df_olm)
    df_olm = make_sex_geno_col(df_olm, sex_col="Sex")
    data["olm"] = df_olm

    return data


# ─────────────────────────────────────────────────────────────────────────────
# FIGURE: SUPPLEMENTAL SEX-STRATIFIED FIGURES
# ─────────────────────────────────────────────────────────────────────────────

def plot_sex_dvc(data):
    """Figure: Summed Activity – All Dark Hours, broken down by sex × genotype."""
    print("  → DVC Summed Dark Activity by Sex")
    df = data["dvc_cage"].dropna(subset=["Sum_All_Dark", "SexGeno"])

    fig, ax = plt.subplots(1, 1, figsize=(3.5, 3.5))
    fig.subplots_adjust(left=0.18, right=0.95, top=0.88, bottom=0.22)

    stats_out = plot_sex_bar_scatter(ax, df, "Sum_All_Dark",
                                     "Total Activity (Dark Phase)",
                                     "Summed Activity (m)")
    return fig, stats_out


def plot_sex_fi_midpoint(data):
    """Figure: F-I Midpoint by sex × genotype."""
    print("  → FI Midpoint by Sex")
    df = data["fi_midpoint"].dropna(subset=["FI_Midpoint", "SexGeno"])

    fig, ax = plt.subplots(1, 1, figsize=(3.5, 3.5))
    fig.subplots_adjust(left=0.18, right=0.95, top=0.88, bottom=0.22)

    stats_out = plot_sex_bar_scatter(ax, df, "FI_Midpoint",
                                     "F-I Curve Midpoint",
                                     "F-I Midpoint (pA)")
    return fig, stats_out


# ISI values used in E/I imbalance line plots (matches Supplemental Figure 1)
_EI_ISIS = [300, 100, 50, 25, 10]


def plot_sex_ei_line(ax, df_ei, pathway, title, ylabel="E/I Imbalance Index",
                     add_legend=False):
    """
    Line plot of E/I Imbalance across all ISIs, one line per sex×genotype group.
    Returns per-ISI per-sex Mann-Whitney stats list.
    """
    # Male = triangle marker (dashed line), Female = circle marker (solid line)
    MARKER_MAP = {"WT-M": "^", "WT-F": "o", "I80T/+-M": "^", "I80T/+-F": "o"}

    df_pw = df_ei[df_ei["Pathway"] == pathway].copy()
    stats_out = []

    for group in SEX_GENO_ORDER:
        means, sems = [], []
        for isi in _EI_ISIS:
            sub = df_pw[(df_pw["SexGeno"] == group) & (df_pw["ISI"] == isi)]["E_I_Imbalance"].dropna()
            means.append(sub.mean() if len(sub) > 0 else np.nan)
            sems.append(sub.sem()  if len(sub) > 1 else 0.0)
        x      = np.arange(len(_EI_ISIS))
        color  = SEX_GENO_COLORS.get(group, "gray")
        ls     = SEX_GENO_LINES.get(group, "-")
        marker = MARKER_MAP.get(group, "o")
        n_vals = [len(df_pw[(df_pw["SexGeno"] == group) & (df_pw["ISI"] == isi)]["E_I_Imbalance"].dropna())
                  for isi in _EI_ISIS]
        n_str  = f"{min(n_vals)}–{max(n_vals)}" if min(n_vals) != max(n_vals) else str(min(n_vals))
        ax.errorbar(x, means, yerr=sems, marker=marker, markersize=3.5, linewidth=1.0,
                    linestyle=ls, capsize=2, capthick=0.5,
                    color=color, label=f"{group} (n={n_str})")

    # Per-ISI Mann-Whitney stats for each sex
    for isi_idx, isi in enumerate(_EI_ISIS):
        for comp, sex_label in [("WT-M vs I80T/+-M", "Male"), ("WT-F vs I80T/+-F", "Female")]:
            geno_a, geno_b = comp.split(" vs ")
            a = df_pw[(df_pw["SexGeno"] == geno_a) & (df_pw["ISI"] == isi)]["E_I_Imbalance"].dropna().values
            b = df_pw[(df_pw["SexGeno"] == geno_b) & (df_pw["ISI"] == isi)]["E_I_Imbalance"].dropna().values
            u, p = mann_whitney(a, b)
            stats_out.append(dict(Pathway=pathway, ISI=isi, Comparison=comp,
                                  SexLabel=sex_label, U=u, P=p, Sig=p_to_sig(p)))

    ax.set_xticks(np.arange(len(_EI_ISIS)))
    ax.set_xticklabels([str(v) for v in _EI_ISIS], fontsize=7)
    ax.set_xlabel("ISI (ms)", fontsize=7)
    ax.set_ylabel(ylabel, fontsize=7)
    ax.set_title(title, fontsize=8)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_ylim(0.2, 1.0)
    ax.set_yticks([0.2, 0.4, 0.6, 0.8, 1.0])

    if add_legend:
        ax.legend(frameon=False, fontsize=5.5, loc="lower right")

    return stats_out


def plot_sex_ei_imbalance(data):
    """
    Figure: E/I Imbalance across all ISIs for all 3 pathways,
    broken down by sex × genotype (4 lines per panel).
    """
    print("  → E/I Imbalance by Sex (all pathways, all ISIs)")
    df_ei = data["ei"]

    pathways_info = [
        ("Perforant",            "ECIII (Perforant)"),
        ("Schaffer",             "CA3 Apical (Schaffer)"),
        ("Basal_Stratum_Oriens", "CA3 Basal"),
    ]

    fig, axes = plt.subplots(1, 3, figsize=(7.5, 2.8))
    fig.subplots_adjust(wspace=0.45, left=0.10, right=0.97, top=0.85, bottom=0.25)

    all_stats = []
    for col_idx, (ax, (pw, pw_label)) in enumerate(zip(axes, pathways_info)):
        s = plot_sex_ei_line(ax, df_ei, pw, pw_label,
                              ylabel="E/I Imbalance Index" if col_idx == 0 else "",
                              add_legend=(col_idx == 2))
        all_stats.extend(s)

    add_subplot_label(axes[0], "A")
    add_subplot_label(axes[1], "B")
    add_subplot_label(axes[2], "C")

    return fig, all_stats


def plot_sex_plateau(data):
    """Figure: Plateau Area, Gabazine condition, by sex × genotype."""
    print("  → Plateau Area by Sex")
    df = data["plateau"]
    # Use Gabazine_Only condition (Condition_Code==1)
    df_gab = df[df["Condition"].str.contains("Gabazine", na=False, case=False)].copy()
    df_gab = df_gab.dropna(subset=["Plateau_Area", "SexGeno"])

    fig, ax = plt.subplots(1, 1, figsize=(3.5, 3.5))
    fig.subplots_adjust(left=0.18, right=0.95, top=0.88, bottom=0.22)

    stats_out = plot_sex_bar_scatter(ax, df_gab, "Plateau_Area",
                                     "Plateau Area (Gabazine)",
                                     "Plateau Area (mV·s)")
    return fig, stats_out


def plot_sex_open_field(data):
    """Figure: Open Field Locomotion + Anxiety by sex × genotype (side by side)."""
    print("  → Open Field by Sex")
    df_loc = data["of_loc"].dropna(subset=["Distance (m)", "SexGeno"])
    df_anx = data["of_anx"].dropna(subset=["Center_Outer_Time_Ratio", "SexGeno"])

    fig, axes = plt.subplots(1, 2, figsize=(7.0, 3.8))
    fig.subplots_adjust(wspace=0.45, left=0.10, right=0.97, top=0.88, bottom=0.22)

    stats_loc = plot_sex_bar_scatter(axes[0], df_loc, "Distance (m)",
                                      "Open Field Locomotion",
                                      "Distance (m)")
    stats_anx = plot_sex_bar_scatter(axes[1], df_anx, "Center_Outer_Time_Ratio",
                                      "Open Field Anxiety",
                                      "Center:Outer Time Ratio")

    add_subplot_label(axes[0], "A")
    add_subplot_label(axes[1], "B")

    return fig, {"location": stats_loc, "anxiety": stats_anx}


def plot_sex_olm(data):
    """Figure: OLM – Testing Discrimination Index by sex × genotype."""
    print("  → OLM by Sex")
    df = data["olm"].dropna(subset=["Testing_DI", "SexGeno"])

    fig, ax = plt.subplots(1, 1, figsize=(3.5, 3.5))
    fig.subplots_adjust(left=0.18, right=0.95, top=0.88, bottom=0.22)

    stats_out = plot_sex_bar_scatter(ax, df, "Testing_DI",
                                     "OLM – Testing Discrimination Index",
                                     "Discrimination Index")
    return fig, stats_out


# ─────────────────────────────────────────────────────────────────────────────
# STATS COLLECTION → Master_Stats_Summary tabs
# ─────────────────────────────────────────────────────────────────────────────

def collect_all_stats(data):
    """Compile a single DataFrame of all sex-stratified stats."""
    all_rows = []

    # ── 1. DVC ──────────────────────────────────────────────────────────────
    df_dvc = data["dvc_cage"].dropna(subset=["Sum_All_Dark", "SexGeno"])
    for comp, sex_label in [("WT-M vs I80T/+-M", "Male"), ("WT-F vs I80T/+-F", "Female")]:
        all_rows.append(build_stats_row(
            "Supplemental – Sex Stratified", "DVC",
            "DVC Summed Dark Activity (m)", "N/A", "All Dark Hours",
            df_dvc, "Sum_All_Dark", sex_label, comp
        ))

    # ── 2. FI Midpoint ──────────────────────────────────────────────────────
    df_fi = data["fi_midpoint"].dropna(subset=["FI_Midpoint", "SexGeno"])
    for comp, sex_label in [("WT-M vs I80T/+-M", "Male"), ("WT-F vs I80T/+-F", "Female")]:
        all_rows.append(build_stats_row(
            "Supplemental – Sex Stratified", "FI Midpoint",
            "F-I Curve Midpoint (pA)", "N/A", "N/A",
            df_fi, "FI_Midpoint", sex_label, comp
        ))

    # ── 3. E/I Imbalance (all ISIs, all 3 pathways) ──────────────────────────
    df_ei = data["ei"].dropna(subset=["E_I_Imbalance", "SexGeno"])
    for pw in ["Perforant", "Schaffer", "Basal_Stratum_Oriens"]:
        df_pw = df_ei[df_ei["Pathway"] == pw]
        for isi in _EI_ISIS:
            df_isi = df_pw[df_pw["ISI"] == isi]
            for comp, sex_label in [("WT-M vs I80T/+-M", "Male"), ("WT-F vs I80T/+-F", "Female")]:
                geno_a, geno_b = comp.split(" vs ")
                grp_a = df_isi[df_isi["SexGeno"] == geno_a]["E_I_Imbalance"].dropna()
                grp_b = df_isi[df_isi["SexGeno"] == geno_b]["E_I_Imbalance"].dropna()
                u, p  = mann_whitney(grp_a.values, grp_b.values)
                def _f(x): return round(float(x), 4) if pd.notna(x) and not np.isnan(float(x)) else np.nan
                all_rows.append(dict(
                    Figure="Supplemental – Sex Stratified",
                    Subpanel=f"E/I Imbalance – {pw}",
                    Metric=f"E/I Imbalance Index [{sex_label}]",
                    Pathway=pw,
                    Condition=f"ISI {isi} ms",
                    WT_Mean=_f(grp_a.mean()), WT_SEM=_f(grp_a.sem()), WT_N=len(grp_a),
                    I80T_Mean=_f(grp_b.mean()), I80T_SEM=_f(grp_b.sem()), I80T_N=len(grp_b),
                    Test_Used="Mann-Whitney U (sex-stratified, per ISI)",
                    Statistic=_f(u), P_Value=_f(p), Significance=p_to_sig(p),
                    Notes=f"Sex-stratified: {comp} | ISI {isi} ms",
                ))

    # ── 4. Plateau Area ─────────────────────────────────────────────────────
    df_plat = data["plateau"]
    df_gab  = df_plat[df_plat["Condition"].str.contains("Gabazine", na=False, case=False)].dropna(
        subset=["Plateau_Area", "SexGeno"])
    for comp, sex_label in [("WT-M vs I80T/+-M", "Male"), ("WT-F vs I80T/+-F", "Female")]:
        all_rows.append(build_stats_row(
            "Supplemental – Sex Stratified", "Plateau Area",
            "Plateau Area (mV·s)", "Both Pathways", "Gabazine condition",
            df_gab, "Plateau_Area", sex_label, comp
        ))

    # ── 5a. Open Field Locomotion ─────────────────────────────────────────
    df_loc = data["of_loc"].dropna(subset=["Distance (m)", "SexGeno"])
    for comp, sex_label in [("WT-M vs I80T/+-M", "Male"), ("WT-F vs I80T/+-F", "Female")]:
        all_rows.append(build_stats_row(
            "Supplemental – Sex Stratified", "Open Field Location",
            "Open Field Distance (m)", "N/A", "Habituation Day 1 Trial 1",
            df_loc, "Distance (m)", sex_label, comp
        ))

    # ── 5b. Open Field Anxiety ────────────────────────────────────────────
    df_anx = data["of_anx"].dropna(subset=["Center_Outer_Time_Ratio", "SexGeno"])
    for comp, sex_label in [("WT-M vs I80T/+-M", "Male"), ("WT-F vs I80T/+-F", "Female")]:
        all_rows.append(build_stats_row(
            "Supplemental – Sex Stratified", "Open Field Anxiety",
            "Open Field Center:Outer Time Ratio", "N/A", "N/A",
            df_anx, "Center_Outer_Time_Ratio", sex_label, comp
        ))

    # ── 6. OLM ────────────────────────────────────────────────────────────
    df_olm = data["olm"].dropna(subset=["Testing_DI", "SexGeno"])
    for comp, sex_label in [("WT-M vs I80T/+-M", "Male"), ("WT-F vs I80T/+-F", "Female")]:
        all_rows.append(build_stats_row(
            "Supplemental – Sex Stratified", "OLM",
            "OLM Testing Discrimination Index", "N/A", "Testing Stage",
            df_olm, "Testing_DI", sex_label, comp
        ))

    return pd.DataFrame(all_rows)


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT STATS TO MASTER EXCEL / CSV
# ─────────────────────────────────────────────────────────────────────────────

COLUMNS = [
    "Figure", "Subpanel", "Metric", "Pathway", "Condition",
    "WT_Mean", "WT_SEM", "WT_N",
    "I80T_Mean", "I80T_SEM", "I80T_N",
    "Test_Used", "Statistic", "P_Value", "Significance",
    "Notes",
]

def export_stats(df_sex_stats):
    """Append sex-stratified stats to Master_Stats_Summary CSV and add tabs to XLSX."""
    # ── CSV: append sex stats block ──────────────────────────────────────────
    if os.path.exists(MASTER_CSV):
        df_existing = pd.read_csv(MASTER_CSV)
        # Remove any previously written sex-stratified rows to avoid duplicates
        mask = df_existing.get("Figure", pd.Series(dtype=str)).str.startswith(
            "Supplemental – Sex", na=False)
        df_existing = df_existing[~mask]
        df_combined = pd.concat([df_existing, df_sex_stats], ignore_index=True)
    else:
        df_combined = df_sex_stats

    df_combined.to_csv(MASTER_CSV, index=False)
    print(f"  ✓ Updated CSV  → {MASTER_CSV} ({len(df_sex_stats)} sex-stratified rows appended)")

    # ── XLSX: add separate tabs ───────────────────────────────────────────────
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font

        # Load existing workbook
        if os.path.exists(MASTER_XLSX):
            wb = load_workbook(MASTER_XLSX)
        else:
            import openpyxl
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # remove default empty sheet

        # Remove stale sex sheets if they exist
        stale = [s for s in wb.sheetnames if "Sex" in s]
        for s in stale:
            del wb[s]

        # Tab 1: All sex-stratified stats combined
        ws_all = wb.create_sheet("Sex Stratified – All")
        ws_all.append(COLUMNS)
        for r in df_sex_stats[COLUMNS].itertuples(index=False):
            ws_all.append(list(r))

        # Tab per metric group
        metric_groups = {
            "Sex - DVC":          lambda d: d["Subpanel"] == "DVC",
            "Sex - FI Midpoint":  lambda d: d["Subpanel"] == "FI Midpoint",
            "Sex - EI Imbalance": lambda d: d["Subpanel"].str.startswith("E/I", na=False),
            "Sex - Plateau Area": lambda d: d["Subpanel"] == "Plateau Area",
            "Sex - Open Field":   lambda d: d["Subpanel"].str.startswith("Open Field", na=False),
            "Sex - OLM":          lambda d: d["Subpanel"] == "OLM",
        }
        for sheet_name, mask_fn in metric_groups.items():
            subset = df_sex_stats[mask_fn(df_sex_stats)]
            if subset.empty:
                continue
            ws = wb.create_sheet(sheet_name[:31])  # Excel sheet name limit
            ws.append(COLUMNS)
            for r in subset[COLUMNS].itertuples(index=False):
                ws.append(list(r))
            # Bold header
            for cell in ws[1]:
                cell.font = Font(bold=True)
            # Auto-width
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

        # Bold header on all-sex tab
        for cell in ws_all[1]:
            cell.font = Font(bold=True)
        for col_cells in ws_all.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws_all.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

        wb.save(MASTER_XLSX)
        print(f"  ✓ Updated XLSX → {MASTER_XLSX} (added {len(metric_groups)+1} sex tabs)")

    except ImportError:
        print("  ⚠ openpyxl not installed – skipping XLSX export. Run: pip install openpyxl")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    setup_publication_style()
    np.random.seed(42)  # reproducible jitter

    print("\n=== Sex-Stratified Figure Generation ===\n")
    print("Loading data …")
    data = load_all_data()

    # ── 1. DVC ──────────────────────────────────────────────────────────────
    print("\n[1] DVC – Summed Activity All Dark Hours")
    fig, _ = plot_sex_dvc(data)
    save_fig_png_svg(fig, "SexStratified_DVC_Summed_Dark_Activity")

    # ── 2. FI Midpoint ───────────────────────────────────────────────────────
    print("\n[2] FI Midpoint")
    fig, _ = plot_sex_fi_midpoint(data)
    save_fig_png_svg(fig, "SexStratified_FI_Midpoint")

    # ── 3. E/I Imbalance ─────────────────────────────────────────────────────
    print("\n[3] E/I Imbalance (all pathways, ISI=300)")
    fig, _ = plot_sex_ei_imbalance(data)
    save_fig_png_svg(fig, "SexStratified_EI_Imbalance")

    # ── 4. Plateau Area ──────────────────────────────────────────────────────
    print("\n[4] Plateau Area")
    fig, _ = plot_sex_plateau(data)
    save_fig_png_svg(fig, "SexStratified_Plateau_Area")

    # ── 5. Open Field ────────────────────────────────────────────────────────
    print("\n[5] Open Field Location + Anxiety")
    fig, _ = plot_sex_open_field(data)
    save_fig_png_svg(fig, "SexStratified_Open_Field")

    # ── 6. OLM ───────────────────────────────────────────────────────────────
    print("\n[6] OLM – Testing Discrimination Index")
    fig, _ = plot_sex_olm(data)
    save_fig_png_svg(fig, "SexStratified_OLM")

    # ── Stats export ──────────────────────────────────────────────────────────
    print("\n[→] Compiling stats …")
    df_sex_stats = collect_all_stats(data)
    # Ensure all COLUMNS present
    for c in COLUMNS:
        if c not in df_sex_stats.columns:
            df_sex_stats[c] = np.nan
    df_sex_stats = df_sex_stats[COLUMNS]

    print(df_sex_stats[["Subpanel", "Metric", "Pathway", "Notes",
                         "WT_Mean", "I80T_Mean", "P_Value", "Significance"]].to_string())

    print("\n[→] Exporting stats to Master_Stats_Summary …")
    export_stats(df_sex_stats)

    print("\n✓ Done. All sex-stratified figures saved to paper_figures/")


if __name__ == "__main__":
    main()
